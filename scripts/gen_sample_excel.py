#!/usr/bin/env python3
"""Generate sample Excel files for filming (Tickets + Visas)"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

OUT_DIR = "/app/public/samples"
os.makedirs(OUT_DIR, exist_ok=True)

# ============================================================
# 1) Tickets sample — 5 rows mixing credit and cash
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = "التذاكر"
ws.sheet_view.rightToLeft = True

headers = [
    "date", "pnr", "passenger_name", "passport_no", "route",
    "carrier_name", "supplier_name", "client_name",
    "currency", "cost", "sale_price", "payment_method", "notes",
]
header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
center = Alignment(horizontal="center", vertical="center")
border = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

rows_t = [
    ["2026-08-07", "IY7823", "محمد أحمد صالح",     "04123456", "SAH-CAI",  "Yemenia", "الخطوط الجوية اليمنية (Yemenia)",   "شركة الأمل للسياحة والحج",  "USD", 220, 260, "credit", ""],
    ["2026-08-07", "SV4567", "سالم عبدالله الحضرمي", "05234567", "ADE-JED",  "Saudia",  "الخطوط السعودية (Saudia)",         "الصندوق الرئيسي",           "SAR", 900, 1100, "cash", "نقد"],
    ["2026-08-08", "IY2201", "أحمد سيف الشعيبي",     "05678921", "SAH-DXB",  "Yemenia", "الخطوط الجوية اليمنية (Yemenia)",   "شركة اليمن الأولى",         "USD", 320, 380, "credit", ""],
    ["2026-08-08", "MS9012", "خالد عبدالكريم",       "04987321", "SAH-CAI",  "MSR",     "الخطوط الجوية اليمنية (Yemenia)",   "الصندوق الرئيسي",           "USD", 180, 220, "cash", "نقد"],
    ["2026-08-09", "SV7788", "بشرى محمد قاسم",       "05412563", "ADE-RUH",  "Saudia",  "الخطوط السعودية (Saudia)",         "مؤسسة الخير للسفر",         "SAR", 1100, 1350, "credit", "طالبة جامعية"],
]
for r, row in enumerate(rows_t, 2):
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.alignment = center
        cell.border = border

# Column widths
widths_t = [12, 10, 22, 14, 12, 12, 32, 28, 8, 10, 10, 12, 15]
for i, w in enumerate(widths_t, 1):
    ws.column_dimensions[chr(64+i) if i <= 26 else 'A'].width = w

wb.save(f"{OUT_DIR}/sample_tickets.xlsx")
print(f"✅ {OUT_DIR}/sample_tickets.xlsx")

# ============================================================
# 2) Visas sample — 4 rows
# ============================================================
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "التأشيرات"
ws2.sheet_view.rightToLeft = True

headers_v = [
    "date", "service_type", "passenger_name", "passport_no", "nationality",
    "entry_date", "expected_exit_date",
    "supplier_name", "client_name",
    "currency", "cost", "sale_price", "payment_method", "notes",
]
for c, h in enumerate(headers_v, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.alignment = center; cell.border = border

rows_v = [
    ["2026-08-07", "تأشيرة عمرة", "فاطمة محمد الشيباني",   "04987654", "يمني",  "2026-08-17", "2026-09-01", "وكيل تأشيرات مكة", "مؤسسة الخير للسفر",         "SAR", 650, 850, "credit", ""],
    ["2026-08-07", "تأشيرة عمرة", "أحمد سيف الشعيبي",     "05678921", "يمني",  "2026-08-20", "2026-09-05", "وكيل تأشيرات مكة", "الصندوق الرئيسي",           "SAR", 650, 850, "cash",   "نقد"],
    ["2026-08-08", "تأشيرة زيارة","نور عبدالكريم",         "05011223", "يمني",  "2026-08-25", "2026-09-15", "وكيل تأشيرات مكة", "شركة الأمل للسياحة والحج",  "SAR", 780, 950, "credit", ""],
    ["2026-08-08", "موافقة أمنية","بندر أحمد الوحش",       "05123999", "يمني",  "2026-09-01", "2026-09-30", "وكيل تأشيرات مكة", "أ. عبدالله الصنعاني",       "USD", 120, 180, "credit", ""],
]
for r, row in enumerate(rows_v, 2):
    for c, v in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=v); cell.alignment = center; cell.border = border

widths_v = [12, 14, 22, 14, 10, 12, 14, 26, 28, 8, 10, 10, 12, 15]
for i, w in enumerate(widths_v, 1):
    ws2.column_dimensions[chr(64+i) if i <= 26 else 'A'+chr(64+i-26)].width = w

wb2.save(f"{OUT_DIR}/sample_visas.xlsx")
print(f"✅ {OUT_DIR}/sample_visas.xlsx")

print("\nAll sample files ready under /app/public/samples/")
print("Download in-app via: NEXT_PUBLIC_BASE_URL/samples/sample_tickets.xlsx")
print("                    NEXT_PUBLIC_BASE_URL/samples/sample_visas.xlsx")
